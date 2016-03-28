'''
Created on Oct 28, 2014

@author: Jim DSouza

Description : TF-IDF portion to select strings that are relevant to property descriptions
'''

import nltk
from nltk.tokenize import sent_tokenize, word_tokenize
from nltk.corpus import stopwords
from nltk.stem.wordnet import WordNetLemmatizer
from nltk import FreqDist

import xlrd
import csv
from openpyxl import load_workbook;

from collections import defaultdict

from time import time
from numpy import zeros
from math import log

from scipy import linalg, dot

from sklearn.preprocessing import Normalizer
from sklearn.cluster import KMeans
from sklearn.decomposition import PCA, TruncatedSVD
from sklearn import metrics

import numpy

from pytagcloud import create_tag_image, make_tags
from pytagcloud.lang.counter import get_tag_counts


def import_file(path):
    t0 = time()
    neg_file = csv.reader(open(path,"r"), delimiter = ",")
    
    tweets = []
    rownum = 0
    for line in neg_file:
        if rownum > 0:
            tweets.append(line[3].strip().upper().encode("utf-8"))
        rownum += 1
    
    print ("Importing file: done in %ds" % (time() - t0));
    return tweets;

### Function to process text by removing stopwords and punctuations ###
def textprocess(tweets):
    t0 = time();
    sent = [];
    
    for tweet in tweets:
        sent.append(sent_tokenize(tweet));
    
    words = [];
    stopwords_eng = stopwords.words('ENGLISH');
    stopwords_eng.remove('AGAINST');
    stopwords_eng.remove('FEW');
    stopwords_eng.remove('NO');
    stopwords_eng.remove('NOR');
    stopwords_eng.remove('NOT');
    stopwords_eng.remove('OFF');
    stopwords_eng.remove('OUT');
    stopwords_eng.remove('OVER');
    stopwords_eng.remove('UNTIL');   
    stopwords_eng.remove('AND');
    stopwords_eng.remove('THE');    
    stopwords_eng.remove('FOR');
    stopwords_eng.remove('HAS');   
    stopwords_eng.remove('HAVE');
    stopwords_eng.remove('HAD');  
    stopwords_eng.remove('THEN');
    stopwords_eng.remove('YET');       
    lemmatizer = WordNetLemmatizer();
    
    for comment in sent:
        temp_words = [];
        for line in comment:
            temp_words.extend(word_tokenize(line));
        temp_words = [word.replace("n't", "not") for word in temp_words];
        temp_words = [word.replace("'", "") for word in temp_words];
        temp_words = [word.replace(",", "") for word in temp_words];
        temp_words = [word.replace("-", "") for word in temp_words];
        temp_words = [word.replace("+", "") for word in temp_words];
        temp_words = [word.replace("/", "") for word in temp_words];
        temp_words = [word.replace("*", "") for word in temp_words];
        temp_words = [word for word in temp_words if not word in stopwords_eng];
        temp_words = [lemmatizer.lemmatize(word) for word in temp_words];
        temp_words = [word for word in temp_words if word != '.'];
        temp_words = [word for word in temp_words if len(word) > 2];
        temp_words = [word.strip() for word in temp_words if not word.isdigit()];
        temp_words = [word for word in temp_words if word != "gudit"];        
        
        words.append(temp_words);
    
    print ("Preproccessing comments: done in %ds" % (time() - t0));
    return words;

### Calculating frequency of words ###
def doc_freq(words):
    t0 = time();
    fd = {}
    for line in words:
        for word in line:
            if word not in fd :
                fd[word] = 1
            else :
                fd[word] += 1
    
    print ("Calculating word frequency: %ds" % (time() - t0));
    return fd;

### Writing words and frequencies to dictionary ###
def write_dict(dict, path):
    writer = csv.writer(open(path, 'wb'));
    writer.writerow(["key", "value"]);
    for key, value in dict.items():
        writer.writerow([key, value]);

### N-grams for words in property descriptions ###
def ngrams(n, words):
    t0 = time();
    ngram_list = [];
    for line in words:
        ngram_list.append([' '.join(item) for item in zip(*[line[i:] for i in range(n)])]);
    
    print ("Creating n grams with %d words. done in %ds" %(n, time()-t0));
    return ngram_list;

### Exporting list to file ###
def write_list(list, path):
    writer = csv.writer(open(path, 'wb'));
    for item in list:
        writer.writerow(item);

def write_txt(list, path):
    file = open(path, "w");
    for item in list:
        file.write(str(item) + "\n");
    file.close();

### Pre-processing functions ###
def dispersion(kmeans, data):
    return numpy.sum(kmeans.transform(data));

def generate_random_dataset(data):
    random_data = numpy.random.uniform(size=data.shape);
    mins = numpy.min(data, axis=0);
    maxs = numpy.max(data, axis=0);
    new_data = numpy.subtract(numpy.multiply(random_data, maxs - mins), mins);
    return new_data;

def createcloud(freq, path):
    tags = make_tags(freq);
    create_tag_image(tags, path, size=(1024, 768), fontname='Lobster');


def preprocessing(tweets):
    t0 = time();
    words1 = textprocess(tweets);
    print ("Text processing. done in %ds" % (time() - t0));
    words2 = ngrams(2, words1);
    words3 = ngrams(3, words1);

    print ("Preprocessing. done in %ds" % (time() - t0));
    
    return words1, words2, words3
    
def createfeatures(freq, words1, words2, words3):
    t0 = time();
    fdist = doc_freq(words1);
    write_dict(fdist, "D:\\Move\\TextMining\\neg_txt_freq1.csv");
    unigrams = [key for key in fdist if fdist[key] > freq];

    fdist = doc_freq(words2);
    write_dict(fdist, "D:\\Move\\TextMining\\neg_txt_freq2.csv");
    bigrams = [key for key in fdist if fdist[key] > freq];
        
    fdist = doc_freq(words3);
    write_dict(fdist, "D:\\Move\\TextMining\\neg_txt_freq3.csv");
    trigrams = [key for key in fdist if fdist[key] > freq];
        
#   features = unigrams + bigrams + trigrams;
    features = unigrams + bigrams + trigrams;
        
    del fdist, unigrams, bigrams, trigrams;

    print ("Feature creation. done in %ds" % (time() - t0));
    
    return features

### Text Document Matrix functions ###
def createTDM(features, words1, words2, words3):
    t0 = time();
    rows, cols = len(words1), len(features);
    TDM = zeros([rows, cols]);

    for i in range(rows):
        for j in range(cols):
            word = features[j];
            if (len(word.split())) == 1:
                for w in words1[i]:
                    if word == w:
                        TDM[i, j] += 1;
            elif (len(word.split())) == 2:
                for w in words2[i]:
                    if word == w:
                        TDM[i, j] += 1;
            else:
                for w in words3[i]:
                    if word == w:
                        TDM[i, j] += 1;
        print ("Creating TDM: %10.2f%% done." % (float(i) * 100 / (rows-1)));
        
    print ("TDM Created. done in %ds" % (time() - t0));
    
    return TDM
    
def TFIDFTransform(TDM):
    t0 = time();
    doc_total = len(TDM);
    rows, cols = TDM.shape;
        
    for row in range(rows):
        wordTotal = reduce(lambda x, y: x+y, TDM[row]);
        for col in range(cols):
            TDM[row][col] = float(TDM[row][col]);
            if TDM[row][col] != 0:
                termDocOcc = getTermDocOcc(col, TDM);
                termFrequency = TDM[row][col] / float(wordTotal);
                inverseDocumentFrequency = log(abs(doc_total / float(termDocOcc)));
                TDM[row][col] = termFrequency * inverseDocumentFrequency;
        print ("TF-IDF Transforming: %10.2f%%" % (float(row) * 100/ (rows - 1)));
        
    print ("TF-IDF Transform. done in %ds" % (time() - t0));
    
    return TDM
    
def getTermDocOcc(col, TDM):
    termDocOcc = 0;
    rows, cols = TDM.shape;
        
    for n in xrange(0, rows):
        if TDM[n][col] > 0:
            termDocOcc += 1;
        
    return termDocOcc;

### SVD Functions ###
def svd(TFIDF):
    t0 = time();
    U, S, Vt = linalg.svd(TFIDF);
        
    print ("SVD. done in %ds" % (time() - t0));
    
    return U, S, Vt

def selSVDMatrices(num_features):
    t0 = time();
    U = U[:, 0:num_features];
    S = S[0:num_features];
    Vt = Vt[0:num_features, :];
        
    print ("Selected SVD features. done in %ds" % (time() - t0));

def featureMatrix():
    t0 = time();
    feature_matrix = dot(dot(U, linalg.diagsvd(S, len(S), len(Vt))), Vt);
        
    print ("Created reconstructed feature matrix. done in %ds" % (time() - t0));
    
def normalize(feature_matrix):
    t0 = time();
    feature_matrix = Normalizer(copy = False).fit_transform(feature_matrix);
        
    print ("Normalized features. done in %ds" % (time() - t0));
    
    return feature_matrix
    
def autosvd(components, TFIDF):
    t0 = time();
    lsa = TruncatedSVD(components, algorithm='randomized', n_iterations=1, random_state=777, tol=0.0001);
    feature_matrix = lsa.fit_transform(TFIDF);
        
    print ("Selected SVD features. done in %ds" % (time() - t0));
    
    return feature_matrix
        
def reducedata(components, verbose):
    t0 = time();
    pca = PCA(n_components=components, whiten=True);
    feature_matrix = pca.fit_transform(TDM);
        
    print ("Feature Reduction using PCA. done in %ds" % (time() - t0));
    if verbose:
        print (pca.explained_variance_ratio_);
    
def fitkmeans(clusters, feature_matrix):
    t0 = time();
    clustering = KMeans(clusters, init='k-means++', max_iter=100, n_init=1, tol=0.0001, precompute_distances=True, 
                        verbose=False, random_state=777);
    kmeans = clustering.fit(feature_matrix);
        
    silhoutte_coeff = metrics.silhouette_score(feature_matrix, kmeans.labels_, metric='euclidean');
    print ("Silhoutte Coefficient: %f" % (silhoutte_coeff));
        
    print ("Clustering completed. done in %ds" % (time() - t0));
    
    return kmeans
    
def readTFIDF(path, tweets):
    t0 = time();
    file = open(path, "r");
    reader = csv.reader(file, delimiter=",");
    rows = len(tweets);
    i = 0;
    for row in reader:
        if reader.line_num == 1:
            cols = len(row);
            TFIDF = zeros([rows, cols]);
        j = 0;
        for col in row:
            TFIDF[i, j] = col;
            j += 1;
        i += 1;
        
    print ("Import TF-IDF features. done in %ds" % (time() - t0));
    
    return TFIDF
    
def getclustfreqs(cluster, wcount, words1, words2, words3, kmeans):
    if wcount == 1:
        words = words1;
    elif wcount == 2:
        words = words2;
    elif wcount == 3:
        words = words3;
    else:
        words = ngrams(wcount, words1);
        
    clus_words = [];
        
    for i, label in enumerate(kmeans.labels_):
        if label == cluster:
            clus_words.append(words[i]);
        
    freq = doc_freq(clus_words);
    freq = [(key, freq[key]) for key in freq.keys()];
    freq = freq[0:25];
        
    return freq;


def read_clusters(path):
    wb = xlrd.open_workbook(path);
    ws = wb.sheet_by_name('Tweet');
    
    num_rows = ws.nrows - 1;
    curr_row = -1;
    
    clusters = [];
    tweets = [];
    
    while curr_row < num_rows:
        curr_row += 1;
        row = ws.row(curr_row);
        clusters.append(ws.cell_value(curr_row, 0));
        tweets.append(ws.cell_value(curr_row, 1));

    return clusters, tweets;

def cluster_count(clusters):
    count = defaultdict(int);
    for cluster in clusters:
        count[cluster] += 1;
    
    return count;

def write_cluster_counts(count, path):
    wb = load_workbook(filename = path);
    ws = wb.get_sheet_by_name("Cluster Counts");
    for i, key in enumerate(count):
        rownum = i+3;
        ws.cell(row=rownum, column=1).value = key;
        ws.cell(row=rownum, column=3).value = count[key];
    wb.save(path);

def main():
    
    t0 = time();
    tweets = [];
    words1 = [];
    words2 = [];
    words3 = [];
    features = [];
    TDM = [];
    U = [];
    S = []
    Vt = [];
    feature_matrix = [];
    kmeans = [];
        
    tweets = import_file("D:\\Move\\EDWData.csv")
    
    words1, words2, words3 = preprocessing(tweets)
    
    # This needs to be run only once - to create the TDIDF data set
    features = createfeatures(2, words1, words2, words3)
    TDM = createTDM(features, words1, words2, words3);
    TDM = TFIDFTransform(TDM);
    write_list(TDM, "D:\\Move\\TextMining\\neg_TFIDF.csv")
    
    TFIDF = readTFIDF("D:\\Move\\TextMining\\neg_TFIDF.csv", tweets);
    
    U, S, Vt = svd(TFIDF)
    write_txt(S, "D:\\Move\\TextMining\\neg_singular.csv");
    
    feature_matrix = autosvd(50, TFIDF)
    feature_matrix = normalize(feature_matrix)
    
    kmeans = fitkmeans(10, feature_matrix)
    
    out_file = csv.writer(open("D:\\Move\\TextMining\\feature_matrix.csv", "wb"))
    out_file.writerows(feature_matrix)
    
    # Create a word cloud - illustrative purposes
    for i in range(10):
        for j in range(3, 4, 1):
            freq = getclustfreqs(i, j, words1, words2, words3, kmeans)
            path = "D:\\Move\\TextMining\\cloud_" + str(i) + "_" + str(j) + ".png"
            createcloud(freq, path)
    
    
    wb = load_workbook(filename = "D:\\Move\\TextMining\\Intermediate.xlsx")
    ws = wb.get_sheet_by_name("Tweet");
    for i, label in enumerate(kmeans.labels_):
        ws.cell(row=i, column=0).value = label
        ws.cell(row=i, column=1).value = tweets[i].encode("utf-8")
    wb.save("D:\\Move\\TextMining\\Intermediate.xlsx")
    
    
    # analysis
    clusters, tweets = read_clusters("D:\\Move\\TextMining\\Intermediate.xlsx");

    count = cluster_count(clusters);
    write_cluster_counts(count, "D:\\Move\\TextMining\\Intermediate.xlsx");
    
    
    print "Time to completion : ", time() - t0

    # cluster 2
    # cluster_comments = get_cluster_comments(2, clusters, tweets);
    # words1, words2, words3, features = create_features(cluster_comments, 3);
    # feature_matrix = feature_reduction(words1, words2, words3, features, 20);
    # kmeans = fitkmeans(feature_matrix, 5);

# labels = [];
# for label in kmeans.labels_:
#     if label == 2:
#         labels.append(0);
#     else:
#         labels.append(label);

# for i in range(5):
#     for j in range(2, 4, 1):
#         freq = getclustfreqs(words1, words2, words3, labels, i, j);
#         path = "D:/Work/Yum/TextMining/cluster/cloud_2" + str(i) + "_" + str(j) + ".png";
#         createcloud(freq, path);

# print (cluster_count(labels));

# cluster 4
# cluster_comments = get_cluster_comments(4, clusters, comments);
# words1, words2, words3, features = create_features(cluster_comments, 5);
# feature_matrix = feature_reduction(words1, words2, words3, features, 10);
# kmeans = fitkmeans(feature_matrix, 4);
# 
# labels = [];
# for label in kmeans.labels_:
#     if label == 1:
#         labels.append(0);
#     else:
#         labels.append(label);
# 
# for i in range(4):
#     for j in range(2, 4, 1):
#         freq = getclustfreqs(words1, words2, words3, labels, i, j);
#         path = "D:/Work/Yum/TextMining/cluster/cloud_4" + str(i) + "_" + str(j) + ".png";
#         createcloud(freq, path);
# 
# print (cluster_count(labels));

# cluster 7
# cluster_comments = get_cluster_comments(7, clusters, comments);
# words1, words2, words3, features = create_features(cluster_comments, 5);
# feature_matrix = feature_reduction(words1, words2, words3, features, 20);
# kmeans = fitkmeans(feature_matrix, 2);
# 
# labels = kmeans.labels_; 
# labels = [];
# for label in kmeans.labels_:
#     if label == 1:
#         labels.append(0);
#     else:
#         labels.append(label);
 
# for i in range(1):
#     for j in range(2, 4, 1):
#         freq = getclustfreqs(words1, words2, words3, labels, i, j);
#         path = "D:/Work/Yum/TextMining/cluster/cloud_7" + str(i) + "_" + str(j) + ".png";
#         createcloud(freq, path);
#  
# print (cluster_count(labels));

# cluster 9
# cluster_comments = get_cluster_comments(9, clusters, comments);
# words1, words2, words3, features = create_features(cluster_comments, 3);
# feature_matrix = feature_reduction(words1, words2, words3, features, 20);
# kmeans = fitkmeans(feature_matrix, 3);
 
# labels = kmeans.labels_; 
# labels = [];
# for label in kmeans.labels_:
#     if label == 1:
#         labels.append(0);
#     else:
#         labels.append(label);
 
# for i in range(3):
#     for j in range(2, 4, 1):
#         freq = getclustfreqs(words1, words2, words3, labels, i, j);
#         path = "D:/Work/Yum/TextMining/cluster/cloud_9" + str(i) + "_" + str(j) + ".png";
#         createcloud(freq, path);
 
# print (cluster_count(labels));
    
if __name__ == '__main__':
    main()